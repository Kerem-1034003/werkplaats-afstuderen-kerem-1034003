import pandas as pd
from openai import OpenAI
import json
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

# Laad de API-sleutel en JSON-categorieën
load_dotenv()
openai_api_key = os.getenv('OPENAI_API_KEY2')

# Initialiseer de OpenAI-client
client = OpenAI(api_key=openai_api_key)

# JSON-categorieën inlezen
with open("../updated_categorie_nested.json", "r") as f:
    categories = json.load(f)

# Laad het productbestand
df = pd.read_excel('../herschreven_excel/simpledeal/homcom/homcom2.1.xlsx')

# Functie om AI de categorie te laten bepalen
def bepaal_categorie(post_title, post_content, json_cats):
    prompt = f"""
Op basis van de volgende productinformatie:
Titel: {post_title}
Beschrijving: {post_content}

Gebruik de JSON-categorieën hieronder om het juiste categoriepad te bepalen voor dit product. Volg hierbij strikt de volgende stappen:
Let goed op de specifieke termen zoals 'fietskar', 'kinderfietskar' 'zonwering' 'parasol' en dergelijke.

1. **Zoek de meest specifieke subsubcategorie**: Analyseer eerst welke subsubcategorie in de JSON het meest passend is voor het product. Maak een keuze die het product het best beschrijft.
2. **Koppel de subsubcategorie aan de bijbehorende hoofdcategorie en subcategorie**: Zodra de subsubcategorie is bepaald, kijk in de JSON welke subcategorie en hoofdcategorie daarbij horen. Gebruik alleen de exacte hiërarchie zoals in de JSON gedefinieerd.
3. **Controleer de structuur**:
   - Het resultaat moet in de vorm zijn: "Hoofdcategorie>Subcategorie>Subsubcategorie".
   - Elk niveau mag slechts één keer voorkomen en moet consistent zijn met de JSON.
   - Het mixen van categorieën of het overslaan van niveaus is niet toegestaan.

4. **Outputregels**:
   - Antwoord uitsluitend met het volledige categoriepad zoals: "Hoofdcategorie>Subcategorie>Subsubcategorie".
   - Voeg geen extra uitleg, tekst of termen toe die niet in de JSON voorkomen.
   - Vermijd Engelse woorden of niet-geldige termen.
   - Voorbeeld: Baby & Kind>Babywereld>Fietskar
   - voorbeeld Tuin>Zonwering>Zonwering
   - voorbeeld Tuin>Parasols>Umbrella Stands

JSON-categorieën:
{json.dumps(json_cats, ensure_ascii=False, indent=2)}

Antwoord alleen met het categoriepad.
"""
    response = client.chat.completions.create(
        model="gpt-3.5-turbo",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.3  # Lagere temperatuur voor consistentie
    )
    categorie_pad = response.choices[0].message.content.strip()
    return categorie_pad

# Functie om het categoriepad te valideren
def valideer_categorie(categorie_pad, json_cats):
    niveaus = categorie_pad.split(">")
    
    # Controleer de hoofdcategorie
    if niveaus[0] not in json_cats:
        return False
    
    # Controleer de subcategorie
    if len(niveaus) > 1 and niveaus[1] not in json_cats[niveaus[0]]:
        return False
    
    # Controleer de subsubcategorie
    if len(niveaus) > 2 and niveaus[2] not in json_cats[niveaus[0]][niveaus[1]]:
        return False
    
    return True

# Itereer door de producten en bepaal de categorie
categorie_paden = []
for index, row in df.iterrows():
    print(f"Processing row {index + 1} of {len(df)}")
    
    post_title = row['post_title']  # Kolomnaam voor de producttitel
    post_content = row['post_content']  # Kolomnaam voor de beschrijving

    poging = 0
    geldig = False
    categorie_pad = None

    while not geldig and poging < 3:  # Maximaal 3 pogingen
        categorie_pad = bepaal_categorie(post_title, post_content, categories)
        geldig = valideer_categorie(categorie_pad, categories)
        poging += 1

    if not geldig:
        print(f"Waarschuwing: Geen valide categorie gevonden voor: {post_title}")
        categorie_pad = "Onbekend>Onbekend>Onbekend"  # Fallbackcategorie

    categorie_paden.append(categorie_pad)

# Voeg de categorieën toe aan de DataFrame
df['tax:product_cat'] = categorie_paden

# Sla het resultaat op in een nieuwe Excel
output_file = "../herschreven_excel/simpledeal/homcom/homcom2.2.xlsx"
df.to_excel(output_file, index=False)

print("Verwerking voltooid! Resultaten zijn opgeslagen in:", output_file)

# Kleur de cellen rood waar categorie "Onbekend>Onbekend>Onbekend" is
from openpyxl import load_workbook
from openpyxl.styles import Font

# Open het Excel-bestand voor opmaak
wb = load_workbook(output_file)
ws = wb.active

# Zoek de kolom met de categorieën
category_column = None
for col in ws.iter_cols(1, ws.max_column):
    if col[0].value == "tax:product_cat":  # Zoek de kolomkop
        category_column = col[0].column
        break

# Controleer de rijen en pas rode kleur toe als categorie "Onbekend>Onbekend>Onbekend" is
if category_column:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # Sla de koprij over
        cell = row[category_column - 1]  # Kolomindex aanpassen naar 0-gebaseerde index
        if cell.value == "Onbekend>Onbekend>Onbekend":
            cell.font = Font(color="FF0000")  # Rode tekstkleur

# Sla het bestand opnieuw op met de opmaak
wb.save("../herschreven_excel/simpledeal/homcom/homcom2.3.xlsx")

print("Verwerking voltooid! Resultaten zijn opgeslagen in:", output_file)